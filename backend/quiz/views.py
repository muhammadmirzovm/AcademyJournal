import io
import random
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import generics, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
from django.db.models import Count
from .models import Topic, Question, Game, Team, GameRound
from .serializers import TopicSerializer, QuestionSerializer, GameSerializer, GameListSerializer

User = get_user_model()

TEAM_NAMES = [
    '🦁 Lions', '🐍 Pythons', '🦊 Foxes', '🦅 Eagles',
    '🐉 Dragons', '🦄 Unicorns', '🐺 Wolves', '🦈 Sharks',
]


class IsTeacher(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ('teacher', 'admin')


# ── Topics ────────────────────────────────────────────────────────────────────

class TopicListCreateView(generics.ListCreateAPIView):
    serializer_class   = TopicSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        qs = Topic.objects.all().select_related('created_by')
        if owner := self.request.query_params.get('owner'):
            qs = qs.filter(created_by_id=owner)
        elif self.request.user.academy_id:
            qs = qs.filter(created_by__academy=self.request.user.academy_id)
        else:
            qs = qs.filter(created_by=self.request.user)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class TopicDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = TopicSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        if self.request.user.role == 'admin':
            return Topic.objects.all()
        return Topic.objects.filter(created_by=self.request.user)


# ── Questions ─────────────────────────────────────────────────────────────────

class QuestionListCreateView(generics.ListCreateAPIView):
    serializer_class   = QuestionSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        qs = Question.objects.all().select_related('topic', 'created_by')
        if owner := self.request.query_params.get('owner'):
            qs = qs.filter(created_by_id=owner)
        if t := self.request.query_params.get('topic'):
            qs = qs.filter(topic_id=t)
        if d := self.request.query_params.get('difficulty'):
            qs = qs.filter(difficulty=d)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class QuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = QuestionSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        if self.request.user.role == 'admin':
            return Question.objects.all()
        return Question.objects.filter(created_by=self.request.user)


# ── Bulk import ───────────────────────────────────────────────────────────────

TEMPLATE_HEADERS = ['Mavzu', 'Savol matni', 'Turi', 'Variant A', 'Variant B', 'Variant C', 'Variant D',
                     "To'g'ri javob", 'Qiyinlik', 'Ball', 'Maslahat']
TEMPLATE_COL_WIDTHS = [18, 40, 12, 16, 16, 16, 16, 14, 10, 8, 24]

DIFFICULTY_MAP = {
    'oson': Question.EASY, 'easy': Question.EASY,
    "o'rta": Question.MEDIUM, 'orta': Question.MEDIUM, 'medium': Question.MEDIUM,
    'qiyin': Question.HARD, 'murakkab': Question.HARD, 'hard': Question.HARD,
}
TRUE_VALUES  = {"to'g'ri", 'togri', 'true', 'ha'}
FALSE_VALUES = {"noto'g'ri", 'notogri', 'false', "yo'q", 'yoq'}


class QuestionTemplateView(APIView):
    """Downloads a blank .xlsx template (+ instructions + filled examples) for
    bulk-importing questions — the easiest way to fill a large question bank."""

    permission_classes = [IsTeacher]

    def _header_row(self, ws):
        fill = PatternFill('solid', fgColor='0D9488')
        bold_white = Font(bold=True, color='FFFFFF')
        for i, h in enumerate(TEMPLATE_HEADERS, start=1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = bold_white
            cell.fill = fill
        for i, w in enumerate(TEMPLATE_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def get(self, request):
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = 'Savollar'
        self._header_row(ws)

        help_ws = wb.create_sheet('Yordam')
        for i, line in enumerate([
            "TURI ustuniga quyidagilardan birini yozing: mcq, true_false, open",
            '',
            'mcq (4 variantli savol):',
            '  - Variant A, B, C, D ustunlarini to\'ldiring',
            "  - To'g'ri javob ustuniga A, B, C yoki D deb yozing",
            '',
            'true_false (to\'g\'ri/noto\'g\'ri savol):',
            '  - Variant ustunlarini bo\'sh qoldiring',
            "  - To'g'ri javob ustuniga TO'G'RI yoki NOTO'G'RI deb yozing",
            '',
            'open (ochiq javob):',
            '  - Variant ustunlarini bo\'sh qoldiring',
            "  - To'g'ri javob ixtiyoriy (namunaviy javob yozishingiz mumkin, yoki bo'sh qoldiring)",
            '',
            "QIYINLIK ustuniga: oson, o'rta yoki qiyin (bo'sh qoldirsangiz 'oson' deb olinadi)",
            "BALL ustuniga: son (bo'sh qoldirsangiz 1 deb olinadi)",
            "MAVZU hali mavjud bo'lmasa, avtomatik yaratiladi — oldindan yaratish shart emas.",
            '',
            "'Savollar' varag'ini to'ldirib, saytga yuklang. Namuna uchun 'Namuna' varag'iga qarang.",
        ], start=1):
            help_ws.cell(row=i, column=1, value=line)
        help_ws.column_dimensions['A'].width = 75

        example_ws = wb.create_sheet('Namuna')
        self._header_row(example_ws)
        for r, row in enumerate([
            ['Grammatika', 'Present Simple qachon ishlatiladi?', 'mcq',
             'Doimiy odat va faktlar uchun', "Hozir sodir bo'layotgan harakat uchun",
             "O'tgan tugallangan harakat uchun", 'Kelasi reja uchun', 'A', 'oson', 1,
             "Odat va faktlarni eslang"],
            ['Grammatika', 'London — Angliya poytaxti.', 'true_false', '', '', '', '',
             "TO'G'RI", 'oson', 1, ''],
            ['Yozish', "O'zingiz haqingizda 3 ta gap yozing.", 'open', '', '', '', '',
             '', "o'rta", 2, ''],
        ], start=2):
            for c, val in enumerate(row, start=1):
                example_ws.cell(row=r, column=c, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="savollar_shabloni.xlsx"'
        return response


class QuestionImportView(APIView):
    """Bulk-creates questions from an uploaded .xlsx (built from the template
    above). Valid rows are imported; invalid rows are skipped and reported
    with their row number and reason, rather than failing the whole file."""

    permission_classes = [IsTeacher]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': "Fayl yuborilmadi."}, status=400)

        try:
            wb = openpyxl.load_workbook(upload, data_only=True)
        except Exception:
            return Response({'detail': "Faylni ochib bo'lmadi — .xlsx formatida ekanligiga ishonch hosil qiling."}, status=400)

        ws = wb['Savollar'] if 'Savollar' in wb.sheetnames else wb.active

        errors = []
        to_create = []
        topic_cache = {}

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c is None or str(c).strip() == '' for c in row):
                continue

            values = list(row) + [None] * (11 - len(row))
            topic_name, text, qtype, opt_a, opt_b, opt_c, opt_d, correct, difficulty, points, hint = values[:11]

            topic_name = str(topic_name).strip() if topic_name else ''
            text       = str(text).strip() if text else ''
            qtype      = str(qtype).strip().lower() if qtype else ''

            row_errors = []
            if not topic_name:
                row_errors.append("Mavzu bo'sh")
            if not text:
                row_errors.append('Savol matni bo\'sh')
            if qtype not in (Question.MCQ, Question.TRUE_FALSE, Question.OPEN):
                row_errors.append(f"Turi noto'g'ri: \"{qtype}\" (mcq / true_false / open bo'lishi kerak)")

            options = None
            correct_answer = ''

            if qtype == Question.MCQ:
                opts = {'a': opt_a, 'b': opt_b, 'c': opt_c, 'd': opt_d}
                missing = [k.upper() for k, v in opts.items() if not v or not str(v).strip()]
                if missing:
                    row_errors.append(f"Variant(lar) bo'sh: {', '.join(missing)}")
                else:
                    options = {k: str(v).strip() for k, v in opts.items()}
                letter = str(correct).strip().lower() if correct else ''
                if letter not in ('a', 'b', 'c', 'd'):
                    row_errors.append(f"To'g'ri javob noto'g'ri: \"{correct}\" (A/B/C/D bo'lishi kerak)")
                else:
                    correct_answer = letter
            elif qtype == Question.TRUE_FALSE:
                val = str(correct).strip().lower() if correct else ''
                if val in TRUE_VALUES:
                    correct_answer = 'true'
                elif val in FALSE_VALUES:
                    correct_answer = 'false'
                else:
                    row_errors.append(f"To'g'ri javob noto'g'ri: \"{correct}\" (TO'G'RI yoki NOTO'G'RI bo'lishi kerak)")
            elif qtype == Question.OPEN:
                correct_answer = str(correct).strip() if correct else ''

            diff_raw = str(difficulty).strip().lower() if difficulty else ''
            if not diff_raw:
                diff = Question.EASY
            else:
                diff = DIFFICULTY_MAP.get(diff_raw)
                if diff is None:
                    row_errors.append(f"Qiyinlik noto'g'ri: \"{difficulty}\" (oson / o'rta / qiyin)")

            pts = 1
            if points not in (None, ''):
                try:
                    pts = int(points)
                    if pts < 1:
                        raise ValueError
                except (ValueError, TypeError):
                    row_errors.append(f"Ball noto'g'ri: \"{points}\"")

            if row_errors:
                errors.append({'row': idx, 'messages': row_errors})
                continue

            cache_key = topic_name.lower()
            topic = topic_cache.get(cache_key)
            if topic is None:
                topic = Topic.objects.filter(name__iexact=topic_name, created_by=request.user).first()
                if topic is None:
                    topic = Topic.objects.create(name=topic_name, created_by=request.user)
                topic_cache[cache_key] = topic

            to_create.append(Question(
                topic=topic, text=text, answer_type=qtype, points=pts, difficulty=diff,
                options=options, correct_answer=correct_answer,
                hint=str(hint).strip() if hint else '', created_by=request.user,
            ))

        created_count = 0
        if to_create:
            Question.objects.bulk_create(to_create)
            created_count = len(to_create)

        ok = created_count > 0 or not errors
        return Response({'created': created_count, 'errors': errors}, status=200 if ok else 400)


# ── Question Banks ─────────────────────────────────────────────────────────────

class QuestionBankListView(APIView):
    permission_classes = [IsTeacher]

    def get(self, request):
        teacher_ids = Question.objects.values_list('created_by_id', flat=True).distinct()
        teachers = User.objects.filter(id__in=teacher_ids).annotate(
            question_count=Count('questions', distinct=True),
            topic_count=Count('topics', distinct=True),
        )
        result = []
        for t in teachers:
            result.append({
                'id':             t.id,
                'name':           t.get_full_name() or t.username,
                'username':       t.username,
                'question_count': t.question_count,
                'topic_count':    t.topic_count,
                'is_me':          t.id == request.user.id,
            })
        result.sort(key=lambda x: (0 if x['is_me'] else 1, x['name'].lower()))
        return Response(result)


# ── Game helpers ──────────────────────────────────────────────────────────────

def get_group_and_check_teacher(group_pk, user):
    from groups.models import Group
    group = get_object_or_404(Group, pk=group_pk)
    is_teacher = group.teacher == user or user.role == 'admin'
    return group, is_teacher


def game_response(game, is_teacher=False):
    """Serialize game including board. Board is a SerializerMethodField so always present.

    correct_answer is stripped from current_question_data for students —
    it's never used for auto-grading (the teacher marks correct/wrong by
    eye), so there's no reason to ship the answer key to student browsers.
    """
    context = {} if is_teacher else {'hide_correct_answer': True}
    return GameSerializer(game, context=context).data


def pick_questions_by_difficulty(topic_id, diff_counts):
    """
    Pick exactly diff_counts['easy'/'medium'/'hard'] questions per difficulty
    from all teachers in the academy.
    Raises ValueError if not enough questions are available.
    """
    chosen = []
    for diff in ('easy', 'medium', 'hard'):
        count = max(0, int(diff_counts.get(diff) or 0))
        if count == 0:
            continue
        pool = list(Question.objects.filter(topic_id=topic_id, difficulty=diff))
        if len(pool) < count:
            raise ValueError(f'{diff}:{topic_id}:{count}:{len(pool)}')
        random.shuffle(pool)
        chosen.extend(pool[:count])
    random.shuffle(chosen)
    return chosen


# ── Games ─────────────────────────────────────────────────────────────────────

class GameListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, group_pk):
        group, _ = get_group_and_check_teacher(group_pk, request.user)
        games = Game.objects.filter(group=group).order_by('-created_at')
        return Response(GameListSerializer(games, many=True).data)

    def post(self, request, group_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        serializer = GameListSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        game = serializer.save(group=group, created_by=request.user)

        topic_difficulty_counts = request.data.get('topic_difficulty_counts', {})
        chosen = []
        try:
            for topic_id, diff_counts in topic_difficulty_counts.items():
                if isinstance(diff_counts, dict):
                    chosen.extend(pick_questions_by_difficulty(topic_id, diff_counts))
        except ValueError as e:
            game.delete()
            parts = str(e).split(':')
            diff, topic_id, requested, available = parts[0], parts[1], parts[2], parts[3]
            return Response({
                'detail': f'Not enough {diff} questions for topic {topic_id}: '
                          f'requested {requested}, available {available}.'
            }, status=400)

        if chosen:
            game.questions.set(chosen)

        return Response(game_response(game, is_teacher), status=201)


class GameDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        return Response(game_response(game, is_teacher))

    def delete(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        game.delete()
        return Response(status=204)


class GameStartView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        if game.status != Game.WAITING:
            return Response({'detail': 'Game already started.'}, status=400)

        students = list(User.objects.filter(memberships__group=group))
        random.shuffle(students)

        game.teams.all().delete()

        if group.is_individual:
            student = students[0] if students else None
            name = (f'{student.first_name} {student.last_name}'.strip() or student.username) if student else 'Student'
            team = Team.objects.create(game=game, name=name)
            if student:
                team.members.add(student)
        else:
            shuffled_names = random.sample(TEAM_NAMES, min(game.team_count, len(TEAM_NAMES)))
            while len(shuffled_names) < game.team_count:
                shuffled_names.append(f'Team {len(shuffled_names) + 1}')
            teams = [Team.objects.create(game=game, name=shuffled_names[i]) for i in range(game.team_count)]
            for i, student in enumerate(students):
                teams[i % game.team_count].members.add(student)

        questions = list(game.questions.all())
        if questions:
            game.double_question = random.choice(questions)
        game.status = Game.ACTIVE
        game.save()
        return Response(game_response(game, is_teacher))


class GamePickView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        if game.status != Game.ACTIVE:
            return Response({'detail': 'Game not active.'}, status=400)

        question = get_object_or_404(Question, pk=request.data.get('question_id'))
        team     = get_object_or_404(Team, pk=request.data.get('team_id'), game=game)

        if game.rounds.filter(question=question).exists():
            return Response({'detail': 'Already answered.'}, status=400)

        game.current_question = question
        game.current_team     = team
        game.save()
        return Response(game_response(game, is_teacher))


class GameAnswerView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)

        question = game.current_question
        team     = game.current_team
        if not question or not team:
            return Response({'detail': 'No active question.'}, status=400)

        correct       = request.data.get('correct', False)
        steal_team_id = request.data.get('steal_team_id')
        partial_pct   = request.data.get('partial_pct')   # 25 | 50 | 75 | 100 | None
        is_double     = game.double_question_id == question.id
        base_points   = question.points * (2 if is_double else 1)

        if steal_team_id:
            steal_team = get_object_or_404(Team, pk=steal_team_id, game=game)
            pts = base_points // 2
            steal_team.score += pts
            steal_team.save()
            GameRound.objects.create(game=game, question=question, picked_by=team,
                stolen_by=steal_team, is_correct=True, is_stolen=True,
                points_awarded=pts, is_double=is_double)
        elif partial_pct is not None:
            pct = max(0, min(100, int(partial_pct)))
            pts = int(base_points * pct / 100)
            team.score += pts
            team.save()
            GameRound.objects.create(game=game, question=question, picked_by=team,
                is_correct=pct > 0, points_awarded=pts, is_double=is_double)
        elif correct:
            team.score += base_points
            team.save()
            GameRound.objects.create(game=game, question=question, picked_by=team,
                is_correct=True, points_awarded=base_points, is_double=is_double)
        else:
            GameRound.objects.create(game=game, question=question, picked_by=team,
                is_correct=False, points_awarded=0, is_double=is_double)

        game.current_question = None
        game.current_team     = None
        game.save()
        return Response(game_response(game, is_teacher))


class GameFinishView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        game.status = Game.FINISHED
        game.current_question = None
        game.current_team     = None
        game.save()
        return Response(game_response(game, is_teacher))


class GameResetView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        game.rounds.all().delete()
        game.teams.all().update(score=0, final_bet=None)
        questions = list(game.questions.all())
        if questions:
            game.double_question = random.choice(questions)
        game.status = Game.ACTIVE
        game.current_question = None
        game.current_team     = None
        game.save()
        return Response(game_response(game, is_teacher))


class GameSwapMembersView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        if game.status != Game.ACTIVE:
            return Response({'detail': 'Game not active.'}, status=400)
        if game.rounds.exists():
            return Response({'detail': 'Cannot swap after game has started.'}, status=400)

        a_id = request.data.get('student_a')
        b_id = request.data.get('student_b')
        team_a = game.teams.filter(members__id=a_id).first()
        team_b = game.teams.filter(members__id=b_id).first()

        if not team_a or not team_b:
            return Response({'detail': 'Student not found in any team.'}, status=400)
        if team_a == team_b:
            return Response({'detail': 'Students are in the same team.'}, status=400)

        team_a.members.remove(a_id)
        team_b.members.remove(b_id)
        team_a.members.add(b_id)
        team_b.members.add(a_id)
        return Response(game_response(game, is_teacher))


class GameReshuffleView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        game = get_object_or_404(Game, pk=game_pk, group=group)
        if game.status != Game.ACTIVE:
            return Response({'detail': 'Game not active.'}, status=400)
        if game.rounds.exists():
            return Response({'detail': 'Cannot reshuffle after game has started.'}, status=400)

        students = list(User.objects.filter(memberships__group=group))
        random.shuffle(students)
        game.teams.all().delete()
        shuffled_names = random.sample(TEAM_NAMES, min(game.team_count, len(TEAM_NAMES)))
        while len(shuffled_names) < game.team_count:
            shuffled_names.append(f'Team {len(shuffled_names) + 1}')
        teams = [Team.objects.create(game=game, name=shuffled_names[i]) for i in range(game.team_count)]
        for i, student in enumerate(students):
            teams[i % game.team_count].members.add(student)
        return Response(game_response(game, is_teacher))


class GameCopyView(APIView):
    permission_classes = [IsTeacher]

    def post(self, request, group_pk, game_pk):
        group, is_teacher = get_group_and_check_teacher(group_pk, request.user)
        if not is_teacher:
            return Response({'detail': 'Not allowed.'}, status=403)
        original = get_object_or_404(Game, pk=game_pk, group=group)
        copy = Game.objects.create(
            group=group,
            name=f'Copy of {original.name}',
            timer_seconds=original.timer_seconds,
            team_count=original.team_count,
            students_per_team=original.students_per_team,
            created_by=request.user,
        )
        copy.questions.set(original.questions.all())
        return Response(GameListSerializer(copy).data, status=201)
