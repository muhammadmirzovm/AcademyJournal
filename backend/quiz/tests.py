import io

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient
from academies.models import Academy
from groups.models import Group, GroupMembership
from quiz.models import Game, Question, Topic

User = get_user_model()


@pytest.fixture
def academy(db):
    return Academy.objects.create(name='Quiz Test Academy', slug='quiz-test-academy')


@pytest.fixture
def teacher(academy):
    return User.objects.create_user(username='q_teacher', password='pass1234', role='teacher', academy=academy)


@pytest.fixture
def student(academy):
    return User.objects.create_user(username='q_student', password='pass1234', role='student', academy=academy)


def _client_for(username):
    client = APIClient()
    res = client.post('/api/auth/login/', {'username': username, 'password': 'pass1234'})
    client.credentials(HTTP_AUTHORIZATION=f'Bearer {res.data["access"]}')
    return client


def _build_workbook(rows, sheet_name='Savollar'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ['Mavzu', 'Savol matni', 'Turi', 'Variant A', 'Variant B', 'Variant C', 'Variant D',
               "To'g'ri javob", 'Qiyinlik', 'Ball', 'Maslahat']
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _upload(rows, filename='questions.xlsx'):
    content = _build_workbook(rows)
    return SimpleUploadedFile(filename, content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@pytest.mark.django_db
def test_template_download_requires_teacher(teacher, student):
    res = _client_for('q_teacher').get('/api/quiz/questions/template/')
    assert res.status_code == 200
    assert res['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    res2 = _client_for('q_student').get('/api/quiz/questions/template/')
    assert res2.status_code == 403


@pytest.mark.django_db
def test_import_creates_all_three_question_types_and_auto_creates_topic(teacher):
    rows = [
        ['Grammatika', 'Present Simple qachon ishlatiladi?', 'mcq', 'Doimiy odat', 'Hozir sodir bo\'lyapti', "O'tgan", 'Kelasi', 'A', 'oson', 1, 'maslahat'],
        ['Grammatika', 'London Angliya poytaxti.', 'true_false', '', '', '', '', "TO'G'RI", 'oson', 1, ''],
        ['Yozish', "O'zingiz haqingizda yozing.", 'open', '', '', '', '', '', "o'rta", 2, ''],
    ]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 200
    assert res.data['created'] == 3
    assert res.data['errors'] == []

    assert Topic.objects.filter(name='Grammatika', created_by__username='q_teacher').count() == 1
    assert Topic.objects.filter(name='Yozish', created_by__username='q_teacher').count() == 1

    mcq = Question.objects.get(text='Present Simple qachon ishlatiladi?')
    assert mcq.answer_type == 'mcq'
    assert mcq.correct_answer == 'a'
    assert mcq.options == {'a': 'Doimiy odat', 'b': "Hozir sodir bo'lyapti", 'c': "O'tgan", 'd': 'Kelasi'}
    assert mcq.difficulty == 'easy'
    assert mcq.points == 1

    tf = Question.objects.get(answer_type='true_false')
    assert tf.correct_answer == 'true'
    assert tf.options is None

    open_q = Question.objects.get(answer_type='open')
    assert open_q.difficulty == 'medium'
    assert open_q.points == 2


@pytest.mark.django_db
def test_import_reuses_existing_topic_case_insensitively(teacher):
    Topic.objects.create(name='Grammatika', created_by=teacher)
    rows = [['grammatika', 'Savol matni', 'true_false', '', '', '', '', 'TRUE', '', '', '']]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 200
    assert res.data['created'] == 1
    assert Topic.objects.filter(created_by=teacher).count() == 1  # not duplicated


@pytest.mark.django_db
def test_import_reports_row_errors_but_still_imports_valid_rows(teacher):
    rows = [
        ['Mavzu', 'Yaxshi savol', 'mcq', 'A variant', 'B variant', 'C variant', 'D variant', 'A', 'oson', 1, ''],
        ['Mavzu', 'Turi xato savol', 'not_a_type', '', '', '', '', '', '', '', ''],
        ['Mavzu', 'MCQ variant yetishmayapti', 'mcq', 'faqat A', '', '', '', 'A', '', '', ''],
        ['', 'Mavzusiz savol', 'open', '', '', '', '', '', '', '', ''],
    ]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 200
    assert res.data['created'] == 1
    assert len(res.data['errors']) == 3
    error_rows = {e['row'] for e in res.data['errors']}
    assert error_rows == {3, 4, 5}


@pytest.mark.django_db
def test_import_invalid_true_false_value_is_rejected(teacher):
    rows = [['Mavzu', 'Savol', 'true_false', '', '', '', '', 'balki', '', '', '']]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 400
    assert res.data['created'] == 0
    assert "To'g'ri javob" in res.data['errors'][0]['messages'][0]


@pytest.mark.django_db
def test_import_blank_difficulty_and_points_use_defaults(teacher):
    rows = [['Mavzu', 'Savol', 'open', '', '', '', '', '', '', '', '']]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 200
    q = Question.objects.get(text='Savol')
    assert q.difficulty == 'easy'
    assert q.points == 1


@pytest.mark.django_db
def test_import_blank_rows_are_skipped_silently(teacher):
    rows = [
        ['Mavzu', 'Haqiqiy savol', 'open', '', '', '', '', '', '', '', ''],
        [None, None, None, None, None, None, None, None, None, None, None],
    ]
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 200
    assert res.data['created'] == 1
    assert res.data['errors'] == []


@pytest.mark.django_db
def test_import_requires_file(teacher):
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {}, format='multipart')
    assert res.status_code == 400


@pytest.mark.django_db
def test_import_rejects_unreadable_file(teacher):
    bad_file = SimpleUploadedFile('notexcel.xlsx', b'this is not a real xlsx file', content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res = _client_for('q_teacher').post('/api/quiz/questions/import/', {'file': bad_file}, format='multipart')
    assert res.status_code == 400


@pytest.mark.django_db
def test_import_requires_teacher(student):
    rows = [['Mavzu', 'Savol', 'open', '', '', '', '', '', '', '', '']]
    res = _client_for('q_student').post('/api/quiz/questions/import/', {'file': _upload(rows)}, format='multipart')
    assert res.status_code == 403


@pytest.mark.django_db
def test_student_does_not_see_correct_answer_but_teacher_does(teacher, student):
    group = Group.objects.create(name='G1', teacher=teacher, class_days=[0, 2, 4])
    GroupMembership.objects.create(group=group, student=student)

    topic = Topic.objects.create(name='T1', created_by=teacher)
    question = Question.objects.create(
        topic=topic, text='2+2=?', answer_type='mcq',
        options={'a': '3', 'b': '4', 'c': '5', 'd': '6'}, correct_answer='b',
        created_by=teacher,
    )
    game = Game.objects.create(
        group=group, name='Game 1', created_by=teacher,
        current_question=question, status=Game.ACTIVE,
    )

    teacher_res = _client_for('q_teacher').get(f'/api/groups/{group.id}/games/{game.id}/')
    assert teacher_res.status_code == 200
    assert teacher_res.data['current_question_data']['correct_answer'] == 'b'

    student_res = _client_for('q_student').get(f'/api/groups/{group.id}/games/{game.id}/')
    assert student_res.status_code == 200
    assert 'correct_answer' not in student_res.data['current_question_data']
    assert student_res.data['current_question_data']['options'] == {'a': '3', 'b': '4', 'c': '5', 'd': '6'}
