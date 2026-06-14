from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Count
from django.http import HttpResponse
from asgiref.sync import async_to_sync
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import Group, GroupMembership, Lesson, Attendance, Score, Journal, CoinTransaction, HomeworkSubmission, Announcement, Exam, ExamResult
from .serializers import (
    GroupSerializer, MemberSerializer, LessonSerializer,
    AttendanceSerializer, BulkAttendanceSerializer,
    ScoreSerializer, BulkScoreSerializer, JournalSerializer,
    HomeworkSubmissionSerializer, AnnouncementSerializer,
    ExamSerializer, ExamResultSerializer,
)


class IsTeacherOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ('teacher', 'admin')


# ── Groups ────────────────────────────────────────────────────────────────────

class GroupListCreateView(generics.ListCreateAPIView):
    serializer_class = GroupSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsTeacherOrAdmin()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return Group.objects.filter(teacher__academy=user.academy)
        if user.role == 'teacher':
            return Group.objects.filter(teacher=user)
        return Group.objects.filter(memberships__student=user)

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class GroupDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = GroupSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        return Group.objects.all()

    def update(self, request, *args, **kwargs):
        group = self.get_object()
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can edit this group.'}, status=403)
        if request.user.role == 'admin' and 'teacher' in request.data:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                new_teacher = User.objects.get(pk=request.data['teacher'], academy=request.user.academy, role='teacher')
                group.teacher = new_teacher
                group.save(update_fields=['teacher'])
            except User.DoesNotExist:
                return Response({'detail': 'Teacher not found in this academy.'}, status=400)
            if list(request.data.keys()) == ['teacher']:
                return Response(GroupSerializer(group, context={'request': request}).data)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        group = self.get_object()
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can delete this group.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class JoinGroupView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        join_key = request.data.get('join_key', '').strip().upper()
        group = get_object_or_404(Group, join_key=join_key)

        if request.user.role == 'teacher':
            return Response({'detail': 'Teachers cannot join groups as students.'}, status=400)

        if GroupMembership.objects.filter(group=group, student=request.user).exists():
            return Response({'detail': 'Already a member.'}, status=400)

        if group.is_individual and group.memberships.exists():
            return Response({'detail': 'Individual group already has a student.'}, status=400)

        GroupMembership.objects.create(group=group, student=request.user)

        from datetime import date
        today = date.today()
        for lesson in group.lessons.filter(date__gte=today):
            Attendance.objects.get_or_create(lesson=lesson, student=request.user, defaults={'present': False})

        return Response(GroupSerializer(group, context={'request': request}).data, status=201)


class GroupMembersView(generics.ListAPIView):
    serializer_class = MemberSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        group = get_object_or_404(Group, pk=self.kwargs['pk'])
        return group.memberships.select_related('student')

    def get_serializer_context(self):
        ctx   = super().get_serializer_context()
        group = get_object_or_404(Group, pk=self.kwargs['pk'])

        memberships  = list(group.memberships.select_related('student').all())
        student_ids  = [m.student.id for m in memberships]

        # ── Comprehension (per-student join-date scoped) ──────────────────────
        comprehension_map = {}
        for membership in memberships:
            join_date     = membership.joined_at.date()
            lessons_since = group.lessons.filter(date__gte=join_date)
            lesson_count  = lessons_since.count()
            if lesson_count == 0:
                comprehension_map[membership.student.id] = None
            else:
                score_sum = (
                    Score.objects
                    .filter(lesson__in=lessons_since, student=membership.student)
                    .aggregate(total=Sum('value'))['total'] or 0
                )
                comprehension_map[membership.student.id] = round(score_sum / (lesson_count * 5) * 100)

        # ── Coins (bulk) ──────────────────────────────────────────────────────
        coin_rows = (
            CoinTransaction.objects
            .filter(group=group, student_id__in=student_ids)
            .values('student')
            .annotate(total=Sum('amount'))
        )
        coin_map = {row['student']: max(0, row['total'] or 0) for row in coin_rows}

        # ── Attendance rate (per-student join-date scoped) ────────────────────
        attendance_map = {}
        for membership in memberships:
            join_date  = membership.joined_at.date()
            lesson_ids = list(group.lessons.filter(date__gte=join_date).values_list('id', flat=True))
            total   = Attendance.objects.filter(lesson_id__in=lesson_ids, student=membership.student).count()
            present = Attendance.objects.filter(lesson_id__in=lesson_ids, student=membership.student, present=True).count()
            attendance_map[membership.student.id] = round(present / total * 100) if total > 0 else None

        from users.models import ParentStudent
        from django.contrib.auth import get_user_model
        _User = get_user_model()
        ctx['parent_set'] = set(
            ParentStudent.objects
            .filter(student_id__in=student_ids)
            .values_list('student_id', flat=True)
        )
        ctx['telegram_set'] = set(
            _User.objects
            .filter(id__in=student_ids, telegram_id__isnull=False)
            .values_list('id', flat=True)
        )

        ctx['comprehension_map'] = comprehension_map
        ctx['coin_map']          = {sid: coin_map.get(sid, 0) for sid in student_ids}
        ctx['attendance_map']    = attendance_map
        return ctx


# ── Lessons ───────────────────────────────────────────────────────────────────

class LessonListCreateView(generics.ListCreateAPIView):
    serializer_class = LessonSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        return Lesson.objects.filter(group_id=self.kwargs['group_pk'])

    def perform_create(self, serializer):
        group = get_object_or_404(Group, pk=self.kwargs['group_pk'])
        if group.teacher != self.request.user and self.request.user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only the teacher or admin can add lessons.')
        lesson = serializer.save(group=group)
        for membership in group.memberships.all():
            Attendance.objects.get_or_create(lesson=lesson, student=membership.student, defaults={'present': False})


class LessonDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = LessonSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        return Lesson.objects.filter(group_id=self.kwargs['group_pk'])

    def update(self, request, *args, **kwargs):
        lesson = self.get_object()
        if lesson.group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can edit lessons.'}, status=403)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        lesson = self.get_object()
        if lesson.group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can delete lessons.'}, status=403)
        return super().destroy(request, *args, **kwargs)


# ── Attendance ────────────────────────────────────────────────────────────────

class AttendanceView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk, lesson_pk):
        lesson  = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        records = Attendance.objects.filter(lesson=lesson).select_related('student')
        return Response(AttendanceSerializer(records, many=True).data)

    def post(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if lesson.group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can mark attendance.'}, status=403)

        serializer = BulkAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        for record in serializer.validated_data['records']:
            Attendance.objects.update_or_create(
                lesson=lesson, student_id=record['student'],
                defaults={'present': record['present']}
            )

        records = Attendance.objects.filter(lesson=lesson).select_related('student')
        return Response(AttendanceSerializer(records, many=True).data)


# ── Scores ────────────────────────────────────────────────────────────────────

class ScoreView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk, lesson_pk):
        lesson  = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        records = Score.objects.filter(lesson=lesson).select_related('student')
        return Response(ScoreSerializer(records, many=True).data)

    def post(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if lesson.group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can enter scores.'}, status=403)

        serializer = BulkScoreSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        for record in serializer.validated_data['records']:
            Score.objects.update_or_create(
                lesson=lesson, student_id=record['student'],
                defaults={'value': int(record['value'])}
            )

        records = Score.objects.filter(lesson=lesson).select_related('student')
        return Response(ScoreSerializer(records, many=True).data)


# ── Journals ──────────────────────────────────────────────────────────────────

class JournalView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if lesson.group.teacher == request.user:
            records = Journal.objects.filter(lesson=lesson).select_related('student')
        else:
            records = Journal.objects.filter(lesson=lesson, student=request.user)
        return Response(JournalSerializer(records, many=True).data)

    def post(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if request.user.role == 'teacher':
            return Response({'detail': 'Teachers do not submit journals.'}, status=400)

        body = request.data.get('body', '').strip()
        if not body:
            return Response({'detail': 'Journal body cannot be empty.'}, status=400)

        journal, _ = Journal.objects.update_or_create(
            lesson=lesson, student=request.user,
            defaults={'body': body}
        )
        return Response(JournalSerializer(journal).data, status=201)


# ── Homework ──────────────────────────────────────────────────────────────────

class HomeworkView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if lesson.group.teacher == request.user:
            submissions = HomeworkSubmission.objects.filter(lesson=lesson).select_related('student')
        else:
            submissions = HomeworkSubmission.objects.filter(lesson=lesson, student=request.user)
        return Response({
            'assignment': lesson.homework,
            'submissions': HomeworkSubmissionSerializer(submissions, many=True).data,
        })

    def post(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)

        if request.user.role == 'teacher':
            if lesson.group.teacher != request.user:
                return Response({'detail': 'Only the lesson teacher can set homework.'}, status=403)
            assignment = request.data.get('assignment', '').strip()
            lesson.homework = assignment
            lesson.save(update_fields=['homework'])
            return Response({'assignment': lesson.homework})

        body = request.data.get('body', '').strip()
        if not body:
            return Response({'detail': 'Submission cannot be empty.'}, status=400)

        submission, _ = HomeworkSubmission.objects.update_or_create(
            lesson=lesson, student=request.user,
            defaults={'body': body}
        )
        return Response(HomeworkSubmissionSerializer(submission).data, status=201)


# ── Membership ────────────────────────────────────────────────────────────────

class AddMemberDirectView(APIView):
    """POST /groups/<pk>/members/add/  { user_id: int }
    Admin or the group's teacher adds a student directly (no join key needed).
    """
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, pk):
        group = get_object_or_404(Group, pk=pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only teacher or admin.'}, status=403)

        from django.contrib.auth import get_user_model
        User = get_user_model()
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'detail': 'user_id required.'}, status=400)

        student = get_object_or_404(User, pk=user_id, role='student')
        if GroupMembership.objects.filter(group=group, student=student).exists():
            return Response({'detail': 'Already a member.'}, status=400)

        if group.is_individual and group.memberships.exists():
            return Response({'detail': 'Individual group already has a student.'}, status=400)

        GroupMembership.objects.create(group=group, student=student)
        return Response({'detail': 'Added.'}, status=201)


class MembershipDetailView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def patch(self, request, pk, member_pk):
        group = get_object_or_404(Group, pk=pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can update membership.'}, status=403)
        membership = get_object_or_404(GroupMembership, pk=member_pk, group=group)

        joined_at = request.data.get('joined_at')
        if joined_at:
            from datetime import datetime, time as time_
            try:
                d = datetime.strptime(str(joined_at)[:10], '%Y-%m-%d')
                GroupMembership.objects.filter(pk=member_pk).update(
                    joined_at=datetime.combine(d.date(), time_.min)
                )
                membership.refresh_from_db()
                # Remove attendance records for lessons before the new join date
                Attendance.objects.filter(
                    lesson__group=group,
                    lesson__date__lt=d.date(),
                    student=membership.student
                ).delete()
            except ValueError:
                return Response({'detail': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

        join_date     = membership.joined_at.date()
        lessons_since = group.lessons.filter(date__gte=join_date)
        lesson_count  = lessons_since.count()
        if lesson_count == 0:
            comprehension = None
        else:
            score_sum = (
                Score.objects.filter(lesson__in=lessons_since, student=membership.student)
                .aggregate(total=Sum('value'))['total'] or 0
            )
            comprehension = round(score_sum / (lesson_count * 5) * 100)

        lesson_ids  = list(lessons_since.values_list('id', flat=True))
        total_att   = Attendance.objects.filter(lesson_id__in=lesson_ids, student=membership.student).count()
        present_att = Attendance.objects.filter(lesson_id__in=lesson_ids, student=membership.student, present=True).count()
        attendance_rate = round(present_att / total_att * 100) if total_att > 0 else None

        data = MemberSerializer(membership).data
        data['comprehension']  = comprehension
        data['attendance_rate'] = attendance_rate
        return Response(data)

    def delete(self, request, pk, member_pk):
        group = get_object_or_404(Group, pk=pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can remove students.'}, status=403)
        membership = get_object_or_404(GroupMembership, pk=member_pk, group=group)
        membership.delete()
        return Response(status=204)


# ── End Lesson ────────────────────────────────────────────────────────────────

class EndLessonView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, group_pk, lesson_pk):
        lesson = get_object_or_404(Lesson, pk=lesson_pk, group_id=group_pk)
        if lesson.group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can end a lesson.'}, status=403)

        from users.models import Notification
        from users.telegram_bot import send_notification

        group      = lesson.group
        attendances = {a.student_id: a.present for a in Attendance.objects.filter(lesson=lesson)}
        scores      = {s.student_id: s.value   for s in Score.objects.filter(lesson=lesson)}

        memberships = (
            GroupMembership.objects
            .filter(group=group)
            .select_related('student')
            .prefetch_related('student__parents__parent')
        )

        for membership in memberships:
            student = membership.student
            present = attendances.get(student.id, False)
            score   = scores.get(student.id)
            name    = f'{student.first_name} {student.last_name}'.strip() or student.username
            scored  = score is not None
            ntype   = 'score' if scored else 'absent'

            # Pick message key variant
            attendance_key = 'present' if present else 'absent'
            score_key      = 'scored'  if scored  else 'unscored'
            student_key    = f'student_{attendance_key}_{score_key}'
            parent_key     = f'parent_{attendance_key}_{score_key}'

            # Student in-app notification body
            if present:
                s_body = f'{"Qatnashdingiz" if not scored else f"Qatnashdingiz · Ball: {score}/5"} — {group.name}'
            else:
                s_body = f'{"Kelmadingiz" if not scored else f"Kelmadingiz · Ball: {score}/5"} — {group.name}'

            Notification.objects.create(
                user=student, type=ntype,
                title=lesson.title, body=s_body,
            )

            if student.telegram_id:
                try:
                    kwargs = dict(lesson=lesson.title, group=group.name)
                    if scored:
                        kwargs['score'] = score
                    async_to_sync(send_notification)(
                        student.telegram_id, student_key, student.telegram_lang or 'uz', **kwargs,
                    )
                except Exception:
                    pass

            # Parent notifications
            for ps in student.parents.select_related('parent').all():
                parent = ps.parent
                if present:
                    p_body = f'{name} · {"Qatnashdi" if not scored else f"Qatnashdi · Ball: {score}/5"} — {group.name}'
                else:
                    p_body = f'{name} · {"Kelmadi" if not scored else f"Kelmadi · Ball: {score}/5"} — {group.name}'

                Notification.objects.create(
                    user=parent, type=ntype,
                    title=lesson.title, body=p_body,
                )
                if parent.telegram_id:
                    try:
                        kwargs = dict(name=name, lesson=lesson.title, group=group.name)
                        if scored:
                            kwargs['score'] = score
                        async_to_sync(send_notification)(
                            parent.telegram_id, parent_key, parent.telegram_lang or 'uz', **kwargs,
                        )
                    except Exception:
                        pass

        # Send homework notification to individual students via DM
        if lesson.homework.strip():
            for membership in memberships:
                student = membership.student
                if student.telegram_id:
                    try:
                        async_to_sync(send_notification)(
                            student.telegram_id, 'hw_notification',
                            student.telegram_lang or 'uz',
                            lesson=lesson.title,
                            group=group.name,
                            homework=lesson.homework,
                        )
                    except Exception:
                        pass

        # Send lesson summary to the linked student Telegram group chat
        if group.telegram_chat_id:
            try:
                import os
                import html as html_module
                from telegram import Bot as TelegramBot

                def e(text):
                    return html_module.escape(str(text))

                lang = group.language or 'uz'
                all_members   = list(memberships)
                total         = len(all_members)
                present_count = sum(1 for m in all_members if attendances.get(m.student.id, False))
                absent_names  = [
                    (f'{m.student.first_name} {m.student.last_name}'.strip() or m.student.username)
                    for m in all_members if not attendances.get(m.student.id, False)
                ]

                if lang == 'ru':
                    all_present_text  = f"✅ Посещаемость: {present_count}/{total} — Все присутствовали!"
                    partial_pres_text = f"✅ Посещаемость: {present_count}/{total} учеников"
                    absent_header     = "❌ <b>Отсутствовали:</b>"
                    homework_header   = "📝 <b>Домашнее задание:</b>"
                else:
                    all_present_text  = f"✅ Davomat: {present_count}/{total} — Hamma qatnashdi!"
                    partial_pres_text = f"✅ Davomat: {present_count}/{total} o'quvchi qatnashdi"
                    absent_header     = "❌ <b>Qatnashmaganlar:</b>"
                    homework_header   = "📝 <b>Uyga vazifa:</b>"

                lines = [f"📚 <b>{e(lesson.title)}</b> — {lesson.date.strftime('%d.%m.%Y')}", ""]
                lines.append(all_present_text if present_count == total else partial_pres_text)

                if absent_names:
                    lines.append("")
                    lines.append(absent_header)
                    for name in absent_names:
                        lines.append(f"• {e(name)}")

                if lesson.homework.strip():
                    lines.append("")
                    lines.append(homework_header)
                    lines.append(e(lesson.homework))

                bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
                if bot_token:
                    bot = TelegramBot(token=bot_token)
                    async_to_sync(bot.send_message)(
                        chat_id=group.telegram_chat_id,
                        text='\n'.join(lines),
                        parse_mode='HTML',
                    )
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f'Group chat send failed: {e}', exc_info=True)

        return Response({'ok': True, 'notified': len(memberships)})


# ── Coins ─────────────────────────────────────────────────────────────────────

class CoinView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, pk):
        group = get_object_or_404(Group, pk=pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can manage coins.'}, status=403)

        student_id = request.data.get('student')
        amount     = request.data.get('amount')
        note       = request.data.get('note', '').strip()

        if not student_id or amount is None:
            return Response({'detail': 'student and amount are required.'}, status=400)

        try:
            amount = int(amount)
        except (ValueError, TypeError):
            return Response({'detail': 'amount must be an integer.'}, status=400)

        if amount == 0:
            return Response({'detail': 'amount cannot be zero.'}, status=400)

        membership = get_object_or_404(GroupMembership, group=group, student_id=student_id)

        current = (
            CoinTransaction.objects
            .filter(group=group, student_id=student_id)
            .aggregate(total=Sum('amount'))['total'] or 0
        )

        # Floor at 0
        if amount < 0:
            amount = max(amount, -current)
        if amount == 0:
            return Response({'balance': current, 'sticker_count': membership.sticker_count, 'sticker_earned': False})

        CoinTransaction.objects.create(group=group, student_id=student_id, amount=amount, note=note)
        new_balance    = current + amount
        threshold      = group.coin_threshold
        stickers_earned = new_balance // threshold

        if stickers_earned > 0:
            membership.sticker_count += stickers_earned
            membership.save()
            CoinTransaction.objects.create(
                group=group, student_id=student_id,
                amount=-(stickers_earned * threshold),
                note=f'{stickers_earned} sticker(s) earned',
            )
            new_balance = new_balance % threshold

        return Response({
            'balance':       new_balance,
            'sticker_count': membership.sticker_count,
            'sticker_earned': stickers_earned > 0,
        })


def _notify_announcement(ann, recipients):
    from users.models import Notification
    from users.telegram_bot import send_notification

    body_preview = (ann.body or '')[:200]
    msg_key = 'announcement_group' if ann.group else 'announcement'
    tg_kwargs = {'title': ann.title, 'body': body_preview or ann.title}
    if ann.group:
        tg_kwargs['group'] = ann.group.name

    Notification.objects.bulk_create([
        Notification(
            user=u,
            type='announcement',
            title=ann.title,
            body=body_preview,
        )
        for u in recipients
    ])

    if ann.is_pinned:
        for u in recipients:
            if u.telegram_id:
                async_to_sync(send_notification)(
                    u.telegram_id, msg_key,
                    u.telegram_lang or 'uz',
                    **tg_kwargs,
                )

        if ann.group and ann.group.telegram_chat_id:
            try:
                import os, html as html_module
                from telegram import Bot as TelegramBot
                e = lambda t: html_module.escape(str(t))
                bot  = TelegramBot(token=os.environ['TELEGRAM_BOT_TOKEN'])
                lang = ann.group.language or 'uz'
                if lang == 'ru':
                    text = f"📢 <b>Объявление</b> — {e(ann.group.name)}\n\n<b>{e(ann.title)}</b>"
                else:
                    text = f"📢 <b>E'lon</b> — {e(ann.group.name)}\n\n<b>{e(ann.title)}</b>"
                if ann.body:
                    text += f"\n\n{e(ann.body[:500])}"
                async_to_sync(bot.send_message)(
                    chat_id=ann.group.telegram_chat_id,
                    text=text,
                    parse_mode='HTML',
                )
            except Exception as ex:
                import logging
                logging.getLogger(__name__).error('Announcement group chat send error: %s', ex)


def _broadcast_academy_ann_to_group_chats(ann, academy):
    import logging
    from users.telegram_bot import send_notification
    active_groups = Group.objects.filter(
        teacher__academy=academy,
        is_graduated=False,
        telegram_chat_id__isnull=False,
    )
    for group in active_groups:
        lang = group.language or 'uz'
        try:
            async_to_sync(send_notification)(
                group.telegram_chat_id,
                'announcement',
                lang,
                title=ann.title,
                body=ann.body[:500] if ann.body else ann.title,
            )
        except Exception as ex:
            logging.getLogger(__name__).error('Broadcast to group chat %s failed: %s', group.id, ex)


class AcademyAnnouncementView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        anns = Announcement.objects.filter(group=None)
        return Response(AnnouncementSerializer(anns, many=True).data)

    def post(self, request):
        if request.user.role != 'admin':
            return Response({'detail': 'Admin only.'}, status=403)
        ser = AnnouncementSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ann = ser.save(author=request.user, group=None)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        # Only notify members and teachers of active (non-graduated) groups
        active_student_ids = set(
            GroupMembership.objects.filter(
                group__teacher__academy=request.user.academy,
                group__is_graduated=False,
            ).values_list('student_id', flat=True)
        )
        active_teacher_ids = set(
            Group.objects.filter(
                teacher__academy=request.user.academy,
                is_graduated=False,
            ).values_list('teacher_id', flat=True)
        )
        recipients = list(
            User.objects.filter(
                id__in=active_student_ids | active_teacher_ids,
                academy=request.user.academy,
            ).exclude(id=request.user.id)
        )
        _notify_announcement(ann, recipients)
        if ann.is_pinned:
            _broadcast_academy_ann_to_group_chats(ann, request.user.academy)
            from users.views import send_push_notification
            recipient_ids = [u.id for u in recipients]
            send_push_notification(recipient_ids, ann.title, ann.body[:200] if ann.body else ann.title)
        return Response(ser.data, status=201)


class GroupAnnouncementView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, pk):
        group = get_object_or_404(Group, pk=pk)
        is_member  = group.memberships.filter(student=request.user).exists()
        is_teacher = group.teacher == request.user
        is_admin   = request.user.role == 'admin'
        if not (is_member or is_teacher or is_admin):
            return Response({'detail': 'No access.'}, status=403)
        return Response(AnnouncementSerializer(group.announcements.all(), many=True).data)

    def post(self, request, pk):
        group = get_object_or_404(Group, pk=pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Teacher or admin only.'}, status=403)
        ser = AnnouncementSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ann = ser.save(author=request.user, group=group)
        from users.models import ParentStudent
        students = [m.student for m in group.memberships.select_related('student').all()]
        parent_ids = ParentStudent.objects.filter(student__in=students).values_list('parent_id', flat=True)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        parents = list(User.objects.filter(id__in=parent_ids))
        recipients = list(set(students + parents) - {request.user})
        _notify_announcement(ann, recipients)
        return Response(ser.data, status=201)


class AnnouncementDeleteView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def delete(self, request, pk):
        ann = get_object_or_404(Announcement, pk=pk)
        if ann.author != request.user and request.user.role != 'admin':
            return Response({'detail': 'No permission.'}, status=403)
        ann.delete()
        return Response(status=204)


# ── Graduate toggle (teacher or admin) ───────────────────────────────────────

class GroupGraduateView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, group_pk):
        group = get_object_or_404(Group, pk=group_pk)
        if group.teacher != request.user and request.user.role != 'admin':
            return Response({'detail': 'Only the teacher or admin can graduate this group.'}, status=403)
        group.is_graduated = not group.is_graduated
        group.save(update_fields=['is_graduated'])
        return Response({'is_graduated': group.is_graduated})


# ── Exam ready toggle (teacher) ───────────────────────────────────────────────

class GroupExamReadyView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, group_pk):
        group = get_object_or_404(Group, pk=group_pk)
        if group.teacher != request.user:
            return Response({'detail': 'Only the teacher of this group can toggle exam readiness.'}, status=403)
        group.exam_ready = not group.exam_ready
        group.save(update_fields=['exam_ready'])

        if group.exam_ready:
            _notify_exam_ready(group, request.user)

        return Response({'exam_ready': group.exam_ready})


def _notify_exam_ready(group, teacher):
    import os, logging
    import urllib.request, urllib.parse, json as _json
    token = os.environ.get('TELEGRAM_BOT_TOKEN')
    if not token:
        return
    teacher_name = f'{teacher.first_name} {teacher.last_name}'.strip() or teacher.username
    from django.contrib.auth import get_user_model
    User = get_user_model()
    admins = User.objects.filter(academy=teacher.academy, role='admin', telegram_id__isnull=False)
    lang = group.language or 'uz'

    # Build schedule string
    DAY_UZ = ['Du', 'Se', 'Ch', 'Pa', 'Ju', 'Sh', 'Ya']
    DAY_RU = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    day_names = DAY_RU if lang == 'ru' else DAY_UZ
    days_str = ', '.join(day_names[d] for d in sorted(group.class_days)) if group.class_days else '—'
    time_str = group.class_time if group.class_time else '—'
    schedule = f"{days_str} — {time_str}"

    if lang == 'ru':
        text = (f"📝 <b>Группа готова к экзамену</b>\n\n"
                f"<b>{group.name}</b>\n"
                f"Учитель: {teacher_name}\n"
                f"Расписание: {schedule}\n\n"
                f"Пожалуйста, создайте экзамен в системе.")
    else:
        text = (f"📝 <b>Guruh imtihonga tayyor</b>\n\n"
                f"<b>{group.name}</b>\n"
                f"O'qituvchi: {teacher_name}\n"
                f"Dars jadvali: {schedule}\n\n"
                f"Iltimos, tizimda imtihon yarating.")
    for admin in admins:
        url  = f'https://api.telegram.org/bot{token}/sendMessage'
        data = urllib.parse.urlencode({'chat_id': admin.telegram_id, 'text': text, 'parse_mode': 'HTML'}).encode()
        try:
            urllib.request.urlopen(urllib.request.Request(url, data=data, method='POST'), timeout=8)
        except Exception as exc:
            logging.getLogger(__name__).error('exam_ready notify error: %s', exc)


# ── Exams (admin creates, everyone reads) ────────────────────────────────────

class ExamListCreateView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk):
        group = get_object_or_404(Group, pk=group_pk)
        is_admin   = request.user.role == 'admin'
        is_teacher = group.teacher == request.user
        is_member  = group.memberships.filter(student=request.user).exists()
        if not (is_admin or is_teacher or is_member):
            return Response({'detail': 'No access.'}, status=403)
        qs        = group.exams.prefetch_related('results__student').order_by('-created_at')
        page      = max(1, int(request.query_params.get('page', 1)))
        page_size = max(1, min(50, int(request.query_params.get('page_size', 10))))
        total     = qs.count()
        pages     = max(1, (total + page_size - 1) // page_size)
        page      = min(page, pages)
        exams     = qs[(page - 1) * page_size : page * page_size]
        return Response({
            'results': ExamSerializer(exams, many=True).data,
            'total':   total,
            'pages':   pages,
            'page':    page,
        })

    def post(self, request, group_pk):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can create exams.'}, status=403)
        group = get_object_or_404(Group, pk=group_pk)
        ser = ExamSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        exam = ser.save(group=group, created_by=request.user, status=Exam.ACTIVE)
        group.exam_ready = False
        group.save(update_fields=['exam_ready'])
        return Response(ExamSerializer(exam).data, status=201)


class ExamDetailView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, group_pk, exam_pk):
        exam = get_object_or_404(Exam, pk=exam_pk, group_id=group_pk)
        return Response(ExamSerializer(exam).data)

    def patch(self, request, group_pk, exam_pk):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin.'}, status=403)
        exam = get_object_or_404(Exam, pk=exam_pk, group_id=group_pk)
        if 'status' in request.data:
            exam.status = request.data['status']
            exam.save(update_fields=['status'])
        return Response(ExamSerializer(exam).data)


class ExamSubmitView(APIView):
    """
    POST /groups/{gid}/exams/{eid}/submit/
    Body: { results: [{ student: id, scores: [0-5, ...] }, ...] }
    """
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, group_pk, exam_pk):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can submit exam results.'}, status=403)
        exam = get_object_or_404(Exam, pk=exam_pk, group_id=group_pk)
        results_data = request.data.get('results', [])

        for entry in results_data:
            student_id = entry.get('student')
            absent     = bool(entry.get('absent', False))
            scores     = [] if absent else entry.get('scores', [])
            comments   = [] if absent else entry.get('comments', [])
            if not absent and len(scores) != exam.question_count:
                return Response({'detail': f'Expected {exam.question_count} scores per student.'}, status=400)
            for s in scores:
                if not (0 <= int(s) <= 5):
                    return Response({'detail': 'Each score must be 0–5.'}, status=400)
            padded_scores   = [int(s) for s in scores] if not absent else [0] * exam.question_count
            padded_comments = list(comments) + [''] * (exam.question_count - len(comments))
            ExamResult.objects.update_or_create(
                exam=exam, student_id=student_id,
                defaults={
                    'absent':   absent,
                    'scores':   padded_scores,
                    'comments': padded_comments[:exam.question_count],
                },
            )

        exam.status = Exam.FINISHED
        exam.save(update_fields=['status'])
        return Response(ExamSerializer(exam).data)


# ── Upcoming exams dashboard ──────────────────────────────────────────────────

class UpcomingExamsView(APIView):
    """
    GET /exams/upcoming/
    Returns:
      - exam_ready_groups: groups waiting for an exam (admin sees all in academy,
        teacher sees their own)
      - active_exams: non-finished exams the user can see
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        user = request.user
        role = user.role

        if role == 'admin':
            ready_groups = Group.objects.filter(
                teacher__academy=user.academy, exam_ready=True
            ).prefetch_related('memberships')
            active_exams = Exam.objects.filter(
                group__teacher__academy=user.academy
            ).exclude(status=Exam.FINISHED).prefetch_related('results__student', 'group')

        elif role == 'teacher':
            ready_groups = Group.objects.filter(
                teacher=user, exam_ready=True
            ).prefetch_related('memberships')
            active_exams = Exam.objects.filter(
                group__teacher=user
            ).exclude(status=Exam.FINISHED).prefetch_related('results__student', 'group')

        else:  # student
            ready_groups = Group.objects.none()
            active_exams = Exam.objects.filter(
                group__memberships__student=user
            ).exclude(status=Exam.FINISHED).prefetch_related('results__student', 'group')

        def group_data(g):
            return {
                'id':           g.id,
                'name':         g.name,
                'member_count': g.memberships.count(),
                'class_days':   g.class_days or [],
                'class_time':   g.class_time or '',
            }

        return Response({
            'exam_ready_groups': [group_data(g) for g in ready_groups],
            'active_exams':      ExamSerializer(active_exams, many=True).data,
        })


# ── Excel Export ───────────────────────────────────────────────────────────────

class ExportExcelView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, pk):
        group = get_object_or_404(Group, pk=pk)
        user  = request.user

        if user.role not in ('teacher', 'admin'):
            return Response({'detail': 'Forbidden'}, status=403)
        if user.role == 'teacher' and group.teacher != user:
            return Response({'detail': 'Forbidden'}, status=403)

        lessons     = list(group.lessons.order_by('date', 'created_at'))
        memberships = list(group.memberships.select_related('student').order_by('student__first_name', 'student__username'))

        if not lessons or not memberships:
            return Response({'detail': 'No data to export'}, status=404)

        lesson_ids  = [l.id for l in lessons]
        student_ids = [m.student_id for m in memberships]

        att_map   = {}
        score_map = {}
        for a in Attendance.objects.filter(lesson_id__in=lesson_ids, student_id__in=student_ids):
            att_map[(a.lesson_id, a.student_id)] = a.present
        for s in Score.objects.filter(lesson_id__in=lesson_ids, student_id__in=student_ids):
            score_map[(s.lesson_id, s.student_id)] = s.value

        wb = openpyxl.Workbook()

        hdr_green  = PatternFill('solid', fgColor='1A5276')
        hdr_blue   = PatternFill('solid', fgColor='1F618D')
        hdr_brown  = PatternFill('solid', fgColor='6E2F19')
        white_bold = Font(bold=True, color='FFFFFF', size=10)
        center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_mid   = Alignment(horizontal='left',   vertical='center')

        def _make_sheet(ws, title_text, col_fill):
            n_lessons = len(lessons)
            last_data_col = get_column_letter(n_lessons + 2)

            ws.merge_cells(f'A1:{last_data_col}1')
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, size=13)
            ws['A1'].alignment = center
            ws.row_dimensions[1].height = 26

            ws['A2'] = '№';         ws['A2'].font = white_bold; ws['A2'].fill = hdr_green; ws['A2'].alignment = center
            ws['B2'] = "O'quvchi";  ws['B2'].font = white_bold; ws['B2'].fill = hdr_green; ws['B2'].alignment = center
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 24
            ws.row_dimensions[2].height = 34

            for j, lesson in enumerate(lessons):
                col = get_column_letter(j + 3)
                cell = ws[f'{col}2']
                cell.value     = lesson.date.strftime('%d.%m')
                cell.font      = white_bold
                cell.fill      = col_fill
                cell.alignment = center
                ws.column_dimensions[col].width = 6

            return n_lessons

        # ── Sheet 1: Davomat ───────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Davomat'
        n = _make_sheet(ws1, f'{group.name}  —  Davomat', hdr_blue)

        jami_col = get_column_letter(n + 3)
        pct_col  = get_column_letter(n + 4)
        for lbl, col in [('Jami', jami_col), ('%', pct_col)]:
            c = ws1[f'{col}2']
            c.value = lbl; c.font = white_bold; c.fill = hdr_green; c.alignment = center
            ws1.column_dimensions[col].width = 8

        for i, ms in enumerate(memberships):
            row     = i + 3
            student = ms.student
            name    = f'{student.first_name} {student.last_name}'.strip() or student.username

            ws1[f'A{row}'] = i + 1;  ws1[f'A{row}'].alignment = center
            ws1[f'B{row}'] = name;    ws1[f'B{row}'].alignment = left_mid

            present_count = 0
            for j, lesson in enumerate(lessons):
                col    = get_column_letter(j + 3)
                state  = att_map.get((lesson.id, student.id))
                if state is True:
                    ws1[f'{col}{row}'].value = '✓'
                    ws1[f'{col}{row}'].font  = Font(color='166534', bold=True)
                    present_count += 1
                elif state is False:
                    ws1[f'{col}{row}'].value = '✗'
                    ws1[f'{col}{row}'].font  = Font(color='991B1B', bold=True)
                else:
                    ws1[f'{col}{row}'].value = '—'
                    ws1[f'{col}{row}'].font  = Font(color='94A3B8')
                ws1[f'{col}{row}'].alignment = center

            pct = round(present_count / n * 100) if n else 0
            ws1[f'{jami_col}{row}'] = f'{present_count}/{n}'; ws1[f'{jami_col}{row}'].alignment = center; ws1[f'{jami_col}{row}'].font = Font(bold=True)
            pct_cell = ws1[f'{pct_col}{row}']
            pct_cell.value     = f'{pct}%'
            pct_cell.alignment = center
            pct_cell.font      = Font(bold=True, color='166534' if pct >= 80 else ('92400E' if pct >= 60 else '991B1B'))

        # ── Sheet 2: Balllar ───────────────────────────────────────────────────
        ws2 = wb.create_sheet('Balllar')
        _make_sheet(ws2, f'{group.name}  —  Balllar', hdr_brown)

        avg_col = get_column_letter(n + 3)
        c = ws2[f'{avg_col}2']
        c.value = "O'rtacha"; c.font = white_bold; c.fill = hdr_green; c.alignment = center
        ws2.column_dimensions[avg_col].width = 10

        for i, ms in enumerate(memberships):
            row     = i + 3
            student = ms.student
            name    = f'{student.first_name} {student.last_name}'.strip() or student.username

            ws2[f'A{row}'] = i + 1;  ws2[f'A{row}'].alignment = center
            ws2[f'B{row}'] = name;    ws2[f'B{row}'].alignment = left_mid

            vals = []
            for j, lesson in enumerate(lessons):
                col = get_column_letter(j + 3)
                val = score_map.get((lesson.id, student.id))
                if val is not None:
                    ws2[f'{col}{row}'].value     = val
                    ws2[f'{col}{row}'].font       = Font(bold=True, color='166534' if val >= 4 else ('92400E' if val >= 3 else '991B1B'))
                    vals.append(val)
                else:
                    ws2[f'{col}{row}'].value = '—'
                    ws2[f'{col}{row}'].font  = Font(color='94A3B8')
                ws2[f'{col}{row}'].alignment = center

            if vals:
                avg = round(sum(vals) / len(vals), 1)
                ws2[f'{avg_col}{row}'].value     = avg
                ws2[f'{avg_col}{row}'].font       = Font(bold=True, color='166534' if avg >= 4 else ('92400E' if avg >= 3 else '991B1B'))
            else:
                ws2[f'{avg_col}{row}'].value = '—'
                ws2[f'{avg_col}{row}'].font  = Font(color='94A3B8')
            ws2[f'{avg_col}{row}'].alignment = center

        # ── Response ───────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = group.name.replace(' ', '_').replace('/', '-')
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_name}_davomat_balllar.xlsx"'
        return response
